"""
分析伺服电机电子齿轮比计算Excel文件
"""
import openpyxl
import os
from pathlib import Path

def find_excel_files():
    """查找包含'电子齿轮'或'齿轮比'的Excel文件"""
    excel_dir = Path("/home/ubuntu/workspace/excel_files")
    if not excel_dir.exists():
        print(f"Excel目录不存在: {excel_dir}")
        return []
    
    files = []
    for file in excel_dir.glob("*.xlsx"):
        if "电子齿轮" in file.name or "齿轮比" in file.name or "伺服" in file.name:
            files.append(file)
    
    return files

def get_sheet_names_with_keyword(workbook, keyword):
    """获取包含关键字的sheet名称"""
    matching_sheets = []
    for sheet_name in workbook.sheetnames:
        if keyword in sheet_name:
            matching_sheets.append(sheet_name)
    return matching_sheets

def analyze_sheet(ws, sheet_name):
    """分析工作表"""
    print(f"\n{'='*60}")
    print(f"工作表: {sheet_name}")
    print(f"{'='*60}")
    
    # 查找参数和公式
    params = {}
    formulas = []
    
    for row in range(1, min(100, ws.max_row + 1)):
        for col in range(1, min(20, ws.max_column + 1)):
            cell = ws.cell(row, col)
            value = cell.value
            
            if value:
                value_str = str(value).strip()
                # 查找参数标签
                if any(keyword in value_str for keyword in ["编码器", "分辨率", "脉冲", "减速比", "导程", "移动距离", "电子齿轮", "齿轮比"]):
                    # 尝试获取相邻单元格的值
                    if col < ws.max_column:
                        next_cell = ws.cell(row, col + 1)
                        if next_cell.value:
                            params[value_str] = next_cell.value
                            print(f"参数: {value_str} = {next_cell.value}")
                
                # 查找公式
                if cell.data_type == 'f' and '=' in str(cell.value):
                    formulas.append((row, col, str(cell.value)))
    
    if formulas:
        print(f"\n找到 {len(formulas)} 个公式:")
        for row, col, formula in formulas[:10]:  # 只显示前10个
            print(f"  行{row}, 列{col}: {formula}")
    
    return params, formulas

def main():
    """主函数"""
    excel_files = find_excel_files()
    
    if not excel_files:
        print("未找到相关Excel文件")
        print("\n将基于标准电子齿轮比计算公式创建工具")
        print("\n标准公式:")
        print("1. 电子齿轮比 = (编码器分辨率 × 机械减速比) / (负载移动距离 / 电机转数)")
        print("2. 脉冲当量 = 负载移动距离 / (编码器分辨率 × 电子齿轮比 × 机械减速比)")
        print("3. 电子齿轮比 = 编码器分辨率 / (脉冲当量 × 机械减速比)")
        return
    
    for excel_file in excel_files:
        print(f"\n{'#'*60}")
        print(f"分析文件: {excel_file.name}")
        print(f"{'#'*60}")
        
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=False)
            
            # 查找包含关键字的sheet
            keyword_sheets = get_sheet_names_with_keyword(wb, "电子齿轮")
            if not keyword_sheets:
                keyword_sheets = get_sheet_names_with_keyword(wb, "齿轮比")
            if not keyword_sheets:
                keyword_sheets = wb.sheetnames[:3]  # 分析前3个sheet
            
            for sheet_name in keyword_sheets:
                ws = wb[sheet_name]
                analyze_sheet(ws, sheet_name)
        
        except Exception as e:
            print(f"分析文件 {excel_file.name} 时出错: {e}")

if __name__ == "__main__":
    main()

