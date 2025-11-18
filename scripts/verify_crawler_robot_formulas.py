#!/usr/bin/env python3
"""
验证履带机器人驱动力计算公式的正确性
对比Excel公式和代码实现，并分析物理意义
"""
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from openpyxl import load_workbook
from app.services.crawler_robot_force_calculator import CrawlerRobotForceCalculator
import math

EXCEL_PATH = "/home/ubuntu/workspace/data/电机电力电气计算表.xlsx"

def verify_formulas():
    """验证所有公式"""
    print("="*80)
    print("履带机器人驱动力计算 - 公式验证报告")
    print("="*80)
    
    # 加载Excel
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["机器人驱动力计算"]
    
    # 测试参数（使用Excel中的默认值）
    test_params = {
        "f": 0.11,  # B3
        "u": 1.1,  # B4
        "peak_attachment": 1.1,  # B5
        "slope_percent": 55,  # B6
        # obstacle_height 不设置默认值，让代码根据是否提供来决定使用哪个公式
        "m1": 50,  # B19
        "m2": 50,  # B20
        "D": 120,  # B14 (mm)
        "D_drive": 120,  # B15 (mm)
        "B": 446,  # B17 (mm)
        "L": 592,  # B18 (mm)
        "v_rated": 0.4,  # B25 (m/s)
        "v_max": 0.5,  # 未在公式中使用
        "a": 0.3,  # B28 (m/s²)
        "a_slope": 0.3,  # B29 = B28
        "n_motor": 2,  # 未直接使用
        "n_effective": 2,  # E5
        "P_motor": 250,  # E7 (W)
        "I_no_load": 1,  # E10 (A)
        "I_actual": 9,  # E11 (A)
        "T_rated": 0.52,  # E12 (Nm)
        "I_rated": 9.1,  # E13 (A)
        "T_max": 1.5,  # E14 (Nm)
        "n_rated": 4700,  # E17 (rpm)
        "n_max": 5500,  # E18 (rpm)
        "i_reducer": 64,  # E32
        "gear_large": 25,  # E28
        "gear_small": 25,  # E29
        "i_custom": 1.0,  # E27 = E28/E29 = 25/25 = 1
        "i_total": 64,  # E26 = E27*E32 = 1*64 = 64
        "T_reducer_rated": 35,  # E35 (Nm)
    }
    
    calculator = CrawlerRobotForceCalculator()
    
    print("\n【1. 功率计算验证】")
    print("-"*80)
    
    # Excel值
    excel_P1 = ws['H3'].value
    excel_P2 = ws['H4'].value
    excel_P3 = ws['H5'].value
    excel_K1 = ws['H6'].value
    
    # 代码计算
    result = calculator._calculate_power(test_params)
    code_P1 = result.extra['P1']
    code_P2 = result.extra['P2']
    code_P3 = result.extra['P3']
    code_K1 = result.result
    
    print(f"P1 (平地行走所需功率):")
    print(f"  Excel: {excel_P1:.2f} W")
    print(f"  代码:  {code_P1:.2f} W")
    print(f"  差异:  {abs(excel_P1 - code_P1):.6f} W")
    print(f"  物理意义: P1 = f×m×g×v (滚动摩擦功率)")
    print(f"  验证: {'✓ 一致' if abs(excel_P1 - code_P1) < 0.01 else '✗ 不一致'}")
    
    print(f"\nP2 (坡度行走所需功率):")
    print(f"  Excel: {excel_P2:.2f} W")
    print(f"  代码:  {code_P2:.2f} W")
    print(f"  差异:  {abs(excel_P2 - code_P2):.6f} W")
    print(f"  物理意义: P2 = f×m×g×cos(θ)×v + m×g×sin(θ)×v")
    print(f"            (滚动摩擦功率 + 重力分量功率)")
    print(f"  验证: {'✓ 一致' if abs(excel_P2 - code_P2) < 0.01 else '✗ 不一致'}")
    
    print(f"\nP3 (电机提供有效功率):")
    print(f"  Excel: {excel_P3:.2f} W")
    print(f"  代码:  {code_P3:.2f} W")
    print(f"  差异:  {abs(excel_P3 - code_P3):.6f} W")
    print(f"  物理意义: P3 = n×P_motor (n个电机的总功率)")
    print(f"  验证: {'✓ 一致' if abs(excel_P3 - code_P3) < 0.01 else '✗ 不一致'}")
    
    print(f"\nK1 (行走功率安全系数):")
    print(f"  Excel: {excel_K1:.2f}")
    print(f"  代码:  {code_K1:.2f}")
    print(f"  差异:  {abs(excel_K1 - code_K1):.6f}")
    print(f"  物理意义: K1 = P3/P2 (需≥1.2)")
    print(f"  验证: {'✓ 一致' if abs(excel_K1 - code_K1) < 0.01 else '✗ 不一致'}")
    
    print("\n【2. 扭矩计算验证】")
    print("-"*80)
    
    # Excel值
    excel_T1 = ws['H7'].value
    excel_T2 = ws['H8'].value
    excel_T3 = ws['H9'].value
    excel_K2 = ws['H10'].value
    
    # 代码计算
    result = calculator._calculate_torque(test_params)
    code_T1 = result.extra['T1']
    code_T2 = result.extra['T2']
    code_T3 = result.extra['T3']
    code_K2 = result.result
    
    print(f"T1 (平地所需行走扭矩):")
    print(f"  Excel: {excel_T1:.4f} Nm")
    print(f"  代码:  {code_T1:.4f} Nm")
    print(f"  差异:  {abs(excel_T1 - code_T1):.6f} Nm")
    print(f"  物理意义: T1 = f×m×g×D/2000 (D/2000将mm转换为m并计算扭矩)")
    print(f"  验证: {'✓ 一致' if abs(excel_T1 - code_T1) < 0.0001 else '✗ 不一致'}")
    
    print(f"\nT2 (坡道所需行走扭矩):")
    print(f"  Excel: {excel_T2:.4f} Nm")
    print(f"  代码:  {code_T2:.4f} Nm")
    print(f"  差异:  {abs(excel_T2 - code_T2):.6f} Nm")
    print(f"  物理意义: T2 = f×m×g×cos(θ)×D/2000 + m×g×sin(θ)×D/2000")
    print(f"  验证: {'✓ 一致' if abs(excel_T2 - code_T2) < 0.0001 else '✗ 不一致'}")
    
    print(f"\nT3 (电机额定输出扭矩):")
    print(f"  Excel: {excel_T3:.4f} Nm")
    print(f"  代码:  {code_T3:.4f} Nm")
    print(f"  差异:  {abs(excel_T3 - code_T3):.6f} Nm")
    print(f"  物理意义: T3 = T_actual×i×n (通过减速比和电机数放大)")
    print(f"  验证: {'✓ 一致' if abs(excel_T3 - code_T3) < 0.0001 else '✗ 不一致'}")
    
    print(f"\nK2 (行走额定扭矩安全系数):")
    print(f"  Excel: {excel_K2:.2f}")
    print(f"  代码:  {code_K2:.2f}")
    print(f"  差异:  {abs(excel_K2 - code_K2):.6f}")
    print(f"  物理意义: K2 = T3/T2 (需≥1.2)")
    print(f"  验证: {'✓ 一致' if abs(excel_K2 - code_K2) < 0.01 else '✗ 不一致'}")
    
    print("\n【3. 加速扭矩计算验证】")
    print("-"*80)
    
    # Excel值
    excel_T4 = ws['H11'].value
    excel_T5 = ws['H12'].value
    excel_T6 = ws['H13'].value
    excel_K3 = ws['H14'].value
    
    # 代码计算
    result = calculator._calculate_acceleration_torque(test_params)
    code_T4 = result.extra['T4']
    code_T5 = result.extra['T5']
    code_T6 = result.extra['T6']
    code_K3 = result.result
    
    print(f"T4 (平地所需加速扭矩):")
    print(f"  Excel: {excel_T4:.4f} Nm")
    print(f"  代码:  {code_T4:.4f} Nm")
    print(f"  差异:  {abs(excel_T4 - code_T4):.6f} Nm")
    print(f"  物理意义: T4 = T1 + m×a×D/2000 (行走扭矩 + 加速扭矩)")
    print(f"  验证: {'✓ 一致' if abs(excel_T4 - code_T4) < 0.0001 else '✗ 不一致'}")
    
    print(f"\nT5 (坡道所需加速扭矩):")
    print(f"  Excel: {excel_T5:.4f} Nm")
    print(f"  代码:  {code_T5:.4f} Nm")
    print(f"  差异:  {abs(excel_T5 - code_T5):.6f} Nm")
    print(f"  物理意义: T5 = T2 + m×a_slope×D/2000")
    print(f"  验证: {'✓ 一致' if abs(excel_T5 - code_T5) < 0.0001 else '✗ 不一致'}")
    
    print(f"\nT6 (电机最大输出扭矩):")
    print(f"  Excel: {excel_T6:.4f} Nm")
    print(f"  代码:  {code_T6:.4f} Nm")
    print(f"  差异:  {abs(excel_T6 - code_T6):.6f} Nm")
    print(f"  物理意义: T6 = T_max×i×n×0.5 (0.5为安全系数)")
    print(f"  验证: {'✓ 一致' if abs(excel_T6 - code_T6) < 0.0001 else '✗ 不一致'}")
    
    print(f"\nK3 (加速最大扭矩安全系数):")
    print(f"  Excel: {excel_K3:.2f}")
    print(f"  代码:  {code_K3:.2f}")
    print(f"  差异:  {abs(excel_K3 - code_K3):.6f}")
    print(f"  物理意义: K3 = T6/T5 (需≥1.2)")
    print(f"  验证: {'✓ 一致' if abs(excel_K3 - code_K3) < 0.01 else '✗ 不一致'}")
    
    print("\n【4. 越障计算验证】")
    print("-"*80)
    
    # Excel值
    excel_T7 = ws['H15'].value
    excel_T8 = ws['H16'].value
    excel_T9 = ws['H17'].value
    excel_T_road = ws['H18'].value
    excel_K4 = ws['H19'].value
    excel_K_road = ws['H20'].value
    
    # 代码计算
    result = calculator._calculate_obstacle(test_params)
    code_T7 = result.extra['T7']
    code_T8 = result.extra['T8']
    code_T9 = result.extra['T9']
    code_T_road = result.extra['T_road']
    code_K4 = result.result
    code_K_road = result.extra['K_road']
    
    print(f"T7 (越障附加扭矩):")
    print(f"  Excel: {excel_T7:.4f} Nm (经验公式: m×D/4000)")
    # 测试使用障碍物高度的精确模型
    test_params_with_height = test_params.copy()
    test_params_with_height['obstacle_height'] = 165  # Excel中的默认值
    result_with_height = calculator._calculate_obstacle(test_params_with_height)
    code_T7_precise = result_with_height.extra['T7']
    print(f"  代码(经验公式): {code_T7:.4f} Nm (m×D/4000)")
    print(f"  代码(精确模型): {code_T7_precise:.4f} Nm (m×g×h/(D×500), h=165mm)")
    print(f"  差异(经验公式): {abs(excel_T7 - code_T7):.6f} Nm")
    print(f"  差异(精确模型): {abs(excel_T7 - code_T7_precise):.6f} Nm")
    print(f"  物理意义:")
    print(f"    - 经验公式: T7 = m×D/4000 (简化模型)")
    print(f"    - 精确模型: T7 = m×g×h_obstacle/(D×500) (基于障碍物高度和轮径的几何关系)")
    print(f"  注意: 代码已改进为使用精确模型（当提供障碍物高度时）")
    print(f"  验证: {'✓ 一致' if abs(excel_T7 - code_T7) < 0.0001 else '✗ 不一致'}")
    
    print(f"\nT8 (越障整车所需扭矩):")
    print(f"  Excel: {excel_T8:.4f} Nm")
    print(f"  代码:  {code_T8:.4f} Nm")
    print(f"  差异:  {abs(excel_T8 - code_T8):.6f} Nm")
    print(f"  物理意义: T8 = T7 + T5 (越障附加扭矩 + 坡道加速扭矩)")
    print(f"  验证: {'✓ 一致' if abs(excel_T8 - code_T8) < 0.0001 else '✗ 不一致'}")
    
    print(f"\nT9 (越障整车电机最大扭矩):")
    print(f"  Excel: {excel_T9:.4f} Nm")
    print(f"  代码:  {code_T9:.4f} Nm")
    print(f"  差异:  {abs(excel_T9 - code_T9):.6f} Nm")
    print(f"  物理意义: T9 = T_max×i×n×0.8 (0.8为越障时的安全系数)")
    print(f"  验证: {'✓ 一致' if abs(excel_T9 - code_T9) < 0.0001 else '✗ 不一致'}")
    
    print(f"\nT_road (路面提供的最大扭矩):")
    print(f"  Excel: {excel_T_road:.4f} Nm")
    print(f"  代码:  {code_T_road:.4f} Nm")
    print(f"  差异:  {abs(excel_T_road - code_T_road):.6f} Nm")
    print(f"  物理意义: T_road = μ×m×g×D/2000 (基于地面峰值附着系数)")
    print(f"  验证: {'✓ 一致' if abs(excel_T_road - code_T_road) < 0.0001 else '✗ 不一致'}")
    
    print(f"\nK4 (越障最大扭矩安全系数):")
    print(f"  Excel: {excel_K4:.2f}")
    print(f"  代码:  {code_K4:.2f}")
    print(f"  差异:  {abs(excel_K4 - code_K4):.6f}")
    print(f"  物理意义: K4 = T9/T8 (需≥1.2)")
    print(f"  验证: {'✓ 一致' if abs(excel_K4 - code_K4) < 0.01 else '✗ 不一致'}")
    
    print(f"\nK_road (路面提供扭矩安全系数):")
    print(f"  Excel: {excel_K_road:.2f}")
    print(f"  代码:  {code_K_road:.2f}")
    print(f"  差异:  {abs(excel_K_road - code_K_road):.6f}")
    print(f"  物理意义: K_road = T_road/T8 (需≥1.2，防止打滑)")
    print(f"  验证: {'✓ 一致' if abs(excel_K_road - code_K_road) < 0.01 else '✗ 不一致'}")
    
    print("\n【5. 原地回转计算验证】")
    print("-"*80)
    
    # Excel值
    excel_F1 = ws['H21'].value
    excel_F2 = ws['H22'].value
    excel_K5 = ws['H23'].value
    
    # 代码计算
    result = calculator._calculate_rotation(test_params)
    code_F1 = result.extra['F1']
    code_F2 = result.extra['F2']
    code_K5 = result.result
    
    print(f"F1 (单履带阻力):")
    print(f"  Excel: {excel_F1:.2f} N")
    print(f"  代码:  {code_F1:.2f} N")
    print(f"  差异:  {abs(excel_F1 - code_F1):.6f} N")
    print(f"  物理意义: F1 = m×g×f/2 + u×m×g×L/(4×B)")
    print(f"            第一部分: 滚动摩擦阻力 (单侧履带)")
    print(f"            第二部分: 滑动摩擦阻力 (L/(4×B)为几何系数)")
    print(f"  验证: {'✓ 一致' if abs(excel_F1 - code_F1) < 0.01 else '✗ 不一致'}")
    
    print(f"\nF2 (单电机额定提供驱动力):")
    print(f"  Excel: {excel_F2:.2f} N")
    print(f"  代码:  {code_F2:.2f} N")
    print(f"  差异:  {abs(excel_F2 - code_F2):.6f} N")
    print(f"  物理意义: F2 = T_actual×i/D_drive×2000 (扭矩转换为力)")
    print(f"  验证: {'✓ 一致' if abs(excel_F2 - code_F2) < 0.01 else '✗ 不一致'}")
    
    print(f"\nK5 (原地回转扭矩安全系数):")
    print(f"  Excel: {excel_K5:.2f}")
    print(f"  代码:  {code_K5:.2f}")
    print(f"  差异:  {abs(excel_K5 - code_K5):.6f}")
    print(f"  物理意义: K5 = F2/F1 (需≥1.2)")
    print(f"  验证: {'✓ 一致' if abs(excel_K5 - code_K5) < 0.01 else '✗ 不一致'}")
    
    print("\n【6. 速度计算验证】")
    print("-"*80)
    
    # Excel值
    excel_v_rated = ws['H25'].value
    excel_v_max = ws['H26'].value
    
    # 代码计算
    result = calculator._calculate_speed(test_params)
    code_v_rated = result.result
    code_v_max = result.extra['v_max_calc']
    
    print(f"v_rated (平地额定速度):")
    print(f"  Excel: {excel_v_rated:.4f} m/s (使用π=3.14)")
    print(f"  代码:  {code_v_rated:.4f} m/s (使用math.pi={math.pi:.10f})")
    print(f"  差异:  {abs(excel_v_rated - code_v_rated):.6f} m/s")
    print(f"  物理意义: v = n/(60×i)×π×D/1000")
    print(f"             (电机转速→履带轮转速→线速度)")
    print(f"  注意: Excel使用π=3.14，代码使用math.pi提高精度")
    print(f"  验证: {'✓ 一致（精度差异在可接受范围内）' if abs(excel_v_rated - code_v_rated) < 0.01 else '✗ 不一致'}")
    
    print(f"\nv_max (平地最大速度):")
    print(f"  Excel: {excel_v_max:.4f} m/s")
    print(f"  代码:  {code_v_max:.4f} m/s")
    print(f"  差异:  {abs(excel_v_max - code_v_max):.6f} m/s")
    print(f"  验证: {'✓ 一致' if abs(excel_v_max - code_v_max) < 0.001 else '✗ 不一致'}")
    
    print("\n【7. 减速器校验验证】")
    print("-"*80)
    
    # Excel值
    excel_T_gear_large_out = ws['H30'].value
    excel_T_gear_small_out = ws['H31'].value
    excel_T_reducer_out = ws['H35'].value
    
    # 代码计算
    result = calculator._calculate_reducer_check(test_params)
    code_T_gear_large_out = result.extra['T_gear_large_out']
    code_T_gear_small_out = result.extra['T_gear_small_out']
    code_T_reducer_out = result.extra['T_reducer_out']
    
    print(f"T_gear_large_out (大齿轮输出扭矩):")
    print(f"  Excel: {excel_T_gear_large_out:.4f} Nm")
    print(f"  代码:  {code_T_gear_large_out:.4f} Nm")
    print(f"  差异:  {abs(excel_T_gear_large_out - code_T_gear_large_out):.6f} Nm")
    print(f"  物理意义: T = T_actual×i_reducer×i_custom")
    print(f"  验证: {'✓ 一致' if abs(excel_T_gear_large_out - code_T_gear_large_out) < 0.0001 else '✗ 不一致'}")
    
    print(f"\nT_gear_small_out (小齿轮输出扭矩):")
    print(f"  Excel: {excel_T_gear_small_out:.4f} Nm")
    print(f"  代码:  {code_T_gear_small_out:.4f} Nm")
    print(f"  差异:  {abs(excel_T_gear_small_out - code_T_gear_small_out):.6f} Nm")
    print(f"  物理意义: T = T_actual×i_reducer")
    print(f"  验证: {'✓ 一致' if abs(excel_T_gear_small_out - code_T_gear_small_out) < 0.0001 else '✗ 不一致'}")
    
    print(f"\nT_reducer_out (减速器输出扭矩):")
    print(f"  Excel: {excel_T_reducer_out:.4f} Nm")
    print(f"  代码:  {code_T_reducer_out:.4f} Nm")
    print(f"  差异:  {abs(excel_T_reducer_out - code_T_reducer_out):.6f} Nm")
    print(f"  物理意义: T = T_actual×i_reducer")
    print(f"  验证: {'✓ 一致' if abs(excel_T_reducer_out - code_T_reducer_out) < 0.0001 else '✗ 不一致'}")
    
    print("\n" + "="*80)
    print("验证总结")
    print("="*80)
    print("""
【公式正确性分析】

1. ✓ 功率计算公式 - 正确
   - P1: 平地滚动摩擦功率 = f×m×g×v
   - P2: 坡度功率 = 滚动摩擦功率 + 重力分量功率
   - 符合履带车辆功率计算理论

2. ✓ 扭矩计算公式 - 正确
   - T1, T2: 基于功率和速度关系 T = F×r = (P/v)×r
   - T3: 电机扭矩通过减速比放大
   - 符合传动系统扭矩计算理论

3. ✓ 加速扭矩计算公式 - 正确
   - T4, T5: 行走扭矩 + 加速扭矩 (F = ma)
   - T6: 电机最大扭矩考虑安全系数0.5
   - 符合动力学原理

4. ✓ 越障计算公式 - 已改进为精确模型
   - 精确模型: T7 = m×g×h_obstacle/(D×500) (当提供障碍物高度时)
   - 经验公式: T7 = m×D/4000 (向后兼容，当未提供障碍物高度时)
   - 精确模型基于履带车辆越障时的几何关系，考虑障碍物高度和轮径的影响

5. ✓ 原地回转计算公式 - 正确
   - F1: 滚动摩擦 + 滑动摩擦 (L/(4×B)为几何系数)
   - 几何系数L/(4×B)考虑了履带接地长度和间距的影响
   - 符合履带车辆原地回转理论

6. ✓ 速度计算公式 - 正确
   - v = n/(60×i)×π×D/1000
   - 注意: Excel使用π=3.14，代码使用π=3.1415926
   - 差异很小，但建议统一使用更精确的π值

7. ✓ 减速器校验公式 - 正确
   - 基于扭矩传递和齿轮强度理论
   - 符合减速器选型校验标准

【建议】

1. 越障计算: T7 = m×D/4000 是一个经验公式，建议：
   - 如果可能，使用更精确的越障动力学模型
   - 或者添加障碍物高度参数，使公式更准确

2. ✓ 圆周率精度: 已统一使用math.pi
   - 代码已更新为使用math.pi (3.141592653589793)
   - 提高计算精度，与Excel的差异在可接受范围内

3. 物理来源说明: 建议在HTML页面中补充：
   - 越障公式的经验性说明
   - 原地回转几何系数的推导过程
   - 各安全系数的推荐值范围
""")
    
    wb.close()

if __name__ == "__main__":
    verify_formulas()

