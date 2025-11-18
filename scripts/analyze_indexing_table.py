#!/usr/bin/env python3
"""
分析"分度盘机构选型计算"工作表
"""
from openpyxl import load_workbook

EXCEL_PATH = "/home/ubuntu/workspace/data/电机电力电气计算表.xlsx"

def main():
    try:
        wb = load_workbook(EXCEL_PATH, data_only=False)
        
        # 查找分度盘相关的工作表
        sheet_names = wb.sheetnames
        print("="*80)
        print("查找分度盘相关工作表")
        print("="*80)
        
        ws = None
        ws_name = None
        
        # 查找可能的工作表名称
        possible_names = ["分度盘", "分度盘机构选型计算", "indexing table", "分度"]
        for name in possible_names:
            if name in sheet_names:
                ws = wb[name]
                ws_name = name
                break
        
        if ws is None:
            print("\n未找到'分度盘'相关的工作表")
            print("\n请检查Excel文件中是否有以下名称的工作表：")
            for name in possible_names:
                print(f"  - {name}")
            print("\n当前所有工作表：")
            for i, name in enumerate(sheet_names, 1):
                print(f"  {i}. {name}")
            return
        
        print(f"\n找到工作表: {ws_name}")
        print("="*80)
        print("分度盘机构选型计算 - 工作表分析报告")
        print("="*80)
        
        # 打印前150行，查找关键信息
        print("\n【工作表内容预览】")
        print("-"*80)
        for row in range(1, min(150, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(25, ws.max_column + 1)):
                cell = ws.cell(row, col)
                if cell.value:
                    value_str = str(cell.value)[:50]
                    row_data.append(f"{cell.coordinate}:{value_str}")
            if row_data:
                print(f"行{row}: {' | '.join(row_data)}")
        
        print("\n【查找公式】")
        print("-"*80)
        formulas_found = []
        for row in range(1, min(200, ws.max_row + 1)):
            for col in range(1, min(30, ws.max_column + 1)):
                cell = ws.cell(row, col)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formulas_found.append((cell.coordinate, cell.value))
                    print(f"{cell.coordinate}: {cell.value}")
        
        if not formulas_found:
            print("未找到公式（可能工作表使用data_only=True读取）")
            print("\n尝试读取计算值...")
            wb_data = load_workbook(EXCEL_PATH, data_only=True)
            ws_data = wb_data[ws_name]
            print("\n【计算值预览】")
            for row in range(1, min(150, ws_data.max_row + 1)):
                for col in range(1, min(20, ws_data.max_column + 1)):
                    cell = ws_data.cell(row, col)
                    if cell.value is not None:
                        print(f"{cell.coordinate}: {cell.value}")
        
        print("\n【查找关键参数和标签】")
        print("-"*80)
        # 查找可能的参数标签
        param_keywords = ["质量", "直径", "半径", "角度", "时间", "转速", "转矩", "惯量", 
                         "效率", "摩擦", "安全系数", "分度", "盘", "工作台", "转台",
                         "m", "M", "D", "R", "r", "θ", "t", "N", "T", "J", "η", "μ", "S",
                         "kg", "m", "s", "rpm", "Nm", "kgm2"]
        found_params = []
        for row in range(1, min(150, ws.max_row + 1)):
            for col in range(1, min(25, ws.max_column + 1)):
                cell = ws.cell(row, col)
                if cell.value:
                    cell_str = str(cell.value)
                    for keyword in param_keywords:
                        if keyword in cell_str:
                            found_params.append((cell.coordinate, cell_str))
                            print(f"{cell.coordinate}: {cell_str}")
                            break
        
        print("\n【查找计算步骤标题】")
        print("-"*80)
        step_keywords = ["计算", "步骤", "1)", "2)", "3)", "4)", "5)", "6)", "转速", "转矩", "惯量", "必须"]
        for row in range(1, min(150, ws.max_row + 1)):
            for col in range(1, min(25, ws.max_column + 1)):
                cell = ws.cell(row, col)
                if cell.value:
                    cell_str = str(cell.value)
                    for keyword in step_keywords:
                        if keyword in cell_str and len(cell_str) < 50:
                            print(f"{cell.coordinate}: {cell_str}")
                            break
        
    except FileNotFoundError:
        print(f"错误: 找不到文件 {EXCEL_PATH}")
    except Exception as e:
        print(f"错误: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()

