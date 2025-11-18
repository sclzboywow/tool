#!/usr/bin/env python3
"""
分析"皮带轮间歇运动"工作表
"""
from openpyxl import load_workbook

EXCEL_PATH = "/home/ubuntu/workspace/data/电机电力电气计算表.xlsx"

def main():
    wb = load_workbook(EXCEL_PATH, data_only=False)
    ws = wb["皮带轮间歇运动"]
    
    print("="*80)
    print("皮带轮间歇运动选型计算 - 工作表分析报告")
    print("="*80)
    
    # 打印前80行，查找关键信息
    print("\n【工作表内容预览】")
    print("-"*80)
    for row in range(1, min(80, ws.max_row + 1)):
        row_data = []
        for col in range(1, min(15, ws.max_column + 1)):
            cell = ws.cell(row, col)
            if cell.value:
                row_data.append(f"{cell.coordinate}:{str(cell.value)[:35]}")
        if row_data:
            print(f"行{row}: {' | '.join(row_data)}")
    
    print("\n【查找公式】")
    print("-"*80)
    for row in range(1, min(100, ws.max_row + 1)):
        for col in range(1, min(20, ws.max_column + 1)):
            cell = ws.cell(row, col)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                print(f"{cell.coordinate}: {cell.value}")

if __name__ == "__main__":
    main()

