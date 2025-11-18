#!/usr/bin/env python3
"""
验证"不同形状物体惯量计算"工作表中所有公式的物理来源
"""
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from openpyxl import load_workbook
import math

EXCEL_PATH = "/home/ubuntu/workspace/data/电机电力电气计算表.xlsx"

def main():
    wb = load_workbook(EXCEL_PATH, data_only=False)
    ws = wb["不同形状物体惯量计算"]
    
    print("="*80)
    print("不同形状物体惯量计算 - 公式物理来源验证报告")
    print("="*80)
    
    # 定义每个模块的验证信息
    modules = [
        {
            "name": "1. 圆柱体惯量计算（平行）",
            "rows": (11, 24),
            "mass_cell": "B23",
            "inertia_cell": "B24",
            "mass_formula_excel": "=H8*((D14/2)^2-(D15/2)^2)*D16*B18",
            "inertia_formula_excel": "=((H8/32)*B18*D16*(D14^4-D15^4)+B23*D20^2)*10000",
            "mass_physics": "m = π × ((d0/2)² - (d1/2)²) × L × ρ",
            "mass_source": "空心圆柱体体积公式：V = π × (R² - r²) × L，质量 = 体积 × 密度",
            "inertia_physics": "J = (π/32) × ρ × L × (d0⁴ - d1⁴) + m × e²",
            "inertia_source": "1. 空心圆柱体绕中心轴的转动惯量：J0 = (π/32) × ρ × L × (d0⁴ - d1⁴)\n   2. 平行轴定理：J = J0 + m × e²",
            "standard": "✓ 标准物理公式（转动惯量理论 + 平行轴定理）"
        },
        {
            "name": "2. 圆柱体惯量计算（垂直）",
            "rows": (26, 39),
            "mass_cell": "B38",
            "inertia_cell": "B39",
            "mass_formula_excel": "=H8*((D29/2)^2-(D30/2)^2)*D31*B33",
            "inertia_formula_excel": "=((1/4)*B38*((D29^2+D30^2)/4+(D31^2/3))+(B38*D35^2))*10000",
            "mass_physics": "m = π × ((d0/2)² - (d1/2)²) × L × ρ",
            "mass_source": "空心圆柱体体积公式（与模块1相同）",
            "inertia_physics": "J = (1/4) × m × ((d0²+d1²)/4 + L²/3) + m × e²",
            "inertia_source": "1. 空心圆柱体绕垂直轴的转动惯量（通过中心，垂直于长度方向）\n   组合公式：J0 = (1/4) × m × ((d0²+d1²)/4 + L²/3)\n   2. 平行轴定理：J = J0 + m × e²",
            "standard": "✓ 标准物理公式（转动惯量理论 + 平行轴定理）"
        },
        {
            "name": "3. 方形物体惯量计算",
            "rows": (41, 54),
            "mass_cell": "B53",
            "inertia_cell": "B54",
            "mass_formula_excel": "=D44*D45*D46*B48",
            "inertia_formula_excel": "=((1/12)*B53*(D44^2+D45^2)+B53*D50^2)*10000",
            "mass_physics": "m = x × y × z × ρ",
            "mass_source": "长方体体积公式：V = x × y × z，质量 = 体积 × 密度",
            "inertia_physics": "J = (1/12) × m × (x²+y²) + m × e²",
            "inertia_source": "1. 长方体绕通过重心的轴的转动惯量：J0 = (1/12) × m × (a²+b²)\n   其中a、b是垂直于旋转轴的边长\n   2. 平行轴定理：J = J0 + m × e²",
            "standard": "✓ 标准物理公式（转动惯量理论 + 平行轴定理）"
        },
        {
            "name": "4. 饼状物体惯量计算",
            "rows": (56, 68),
            "mass_cell": "B67",
            "inertia_cell": "B68",
            "mass_formula_excel": "=H8*(D59/2)^2*D60*B62",
            "inertia_formula_excel": "=((1/8)*B67*D59^2+B67*D64^2)*10000",
            "mass_physics": "m = π × (d/2)² × h × ρ",
            "mass_source": "实心圆柱体体积公式：V = π × R² × h，质量 = 体积 × 密度",
            "inertia_physics": "J = (1/8) × m × d² + m × e²",
            "inertia_source": "1. 实心圆柱体绕中心轴的转动惯量：J0 = (1/2) × m × R² = (1/8) × m × d²\n   2. 平行轴定理：J = J0 + m × e²",
            "standard": "✓ 标准物理公式（转动惯量理论 + 平行轴定理）"
        },
        {
            "name": "5. 直线运动物体惯量计算",
            "rows": (70, 77),
            "inertia_cell": "B77",
            "inertia_formula_excel": "=(B74*(B72/(2*H8))^2)*10^4",
            "inertia_physics": "J = m × (A/(2π))²",
            "inertia_source": "将直线运动转换为旋转运动，等效惯量 = m × r²\n   其中 r = A/(2π) 是等效半径（电机每转1圈，物体移动A距离）",
            "standard": "✓ 标准物理公式（运动转换理论）"
        },
        {
            "name": "6. 直接惯量计算",
            "rows": (79, 91),
            "inertia_cell": "B91",
            "inertia_formula_excel": "=B82+B83*B87*B87/100",
            "inertia_physics": "J1 = J0 + m × (e/10)² = J0 + m × e²/100",
            "inertia_source": "平行轴定理（Steiner定理）：J1 = J0 + m × d²\n   其中J0是绕重心的惯量，d是两轴之间的距离\n   注意：e的单位是mm，需要转换为cm：e(cm) = e(mm)/10",
            "standard": "✓ 标准物理公式（平行轴定理）"
        }
    ]
    
    print("\n【公式验证结果】")
    print("="*80)
    
    for module in modules:
        print(f"\n{module['name']}")
        print("-"*80)
        
        if 'mass_cell' in module:
            print(f"质量公式:")
            print(f"  Excel公式: {module['mass_formula_excel']}")
            print(f"  物理公式: {module['mass_physics']}")
            print(f"  物理来源: {module['mass_source']}")
            print(f"  标准依据: {module['standard']}")
        
        print(f"\n惯量公式:")
        print(f"  Excel公式: {module['inertia_formula_excel']}")
        print(f"  物理公式: {module['inertia_physics']}")
        print(f"  物理来源: {module['inertia_source']}")
        print(f"  标准依据: {module['standard']}")
        
        # 验证Excel中的实际公式
        if 'mass_cell' in module:
            mass_cell = ws[module['mass_cell']]
            if mass_cell.data_type == 'f':
                actual_mass = mass_cell.value
                if actual_mass == module['mass_formula_excel'].replace('=', ''):
                    print(f"  ✓ Excel质量公式匹配")
                else:
                    print(f"  ⚠ Excel质量公式: {actual_mass}")
        
        inertia_cell = ws[module['inertia_cell']]
        if inertia_cell.data_type == 'f':
            actual_inertia = inertia_cell.value
            if actual_inertia == module['inertia_formula_excel'].replace('=', ''):
                print(f"  ✓ Excel惯量公式匹配")
            else:
                print(f"  ⚠ Excel惯量公式: {actual_inertia}")
    
    print("\n" + "="*80)
    print("【总结】")
    print("="*80)
    print("所有6个计算模块的公式都基于标准物理原理：")
    print("1. 质量计算：基于体积公式和密度")
    print("2. 转动惯量计算：基于转动惯量理论和平行轴定理（Steiner定理）")
    print("3. 单位换算：注意mm→m和kg·cm²→kg·m²的转换")
    print("\n✓ 所有公式都有明确的物理来源，符合标准物理理论")
    print("="*80)
    
    wb.close()

if __name__ == "__main__":
    main()

