#!/usr/bin/env python3
"""
分析三个新工具的工作表：
1. 电动机启动端压计算
2. 小车驱动功率
3. 机器人驱动力计算
"""
from openpyxl import load_workbook

EXCEL_PATH = "/home/ubuntu/workspace/data/电机电力电气计算表.xlsx"

def analyze_sheet(ws_name):
    """分析单个工作表"""
    print("="*80)
    print(f"分析工作表: {ws_name}")
    print("="*80)
    
    wb = load_workbook(EXCEL_PATH, data_only=False)
    ws = wb[ws_name]
    
    # 打印前200行，查找关键信息
    print("\n【工作表内容预览】")
    print("-"*80)
    for row in range(1, min(200, ws.max_row + 1)):
        row_data = []
        for col in range(1, min(30, ws.max_column + 1)):
            cell = ws.cell(row, col)
            if cell.value:
                value_str = str(cell.value)[:60]
                row_data.append(f"{cell.coordinate}:{value_str}")
        if row_data:
            print(f"行{row}: {' | '.join(row_data)}")
    
    print("\n【查找公式】")
    print("-"*80)
    formulas_found = []
    for row in range(1, min(200, ws.max_row + 1)):
        for col in range(1, min(30, ws.max_column + 1)):
            cell = ws.cell(row, col)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formulas_found.append((cell.coordinate, cell.value))
                print(f"{cell.coordinate}: {cell.value}")
    
    if not formulas_found:
        print("未找到公式（可能工作表使用data_only=True读取）")
        print("\n尝试读取计算值...")
        wb_data = load_workbook(EXCEL_PATH, data_only=True)
        ws_data = wb_data[ws_name]
        print("\n【计算值预览】")
        for row in range(1, min(200, ws_data.max_row + 1)):
            for col in range(1, min(25, ws_data.max_column + 1)):
                cell = ws_data.cell(row, col)
                if cell.value is not None:
                    print(f"{cell.coordinate}: {cell.value}")
    
    print("\n" + "="*80 + "\n")

def main():
    sheets = ["电动机启动端压计算", "小车驱动功率", "机器人驱动力计算"]
    for sheet_name in sheets:
        try:
            analyze_sheet(sheet_name)
        except Exception as e:
            print(f"分析工作表 {sheet_name} 时出错: {e}\n")

if __name__ == "__main__":
    main()

