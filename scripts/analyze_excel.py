#!/usr/bin/env python3
"""分析Excel文件结构"""
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

try:
    from openpyxl import load_workbook
    import json
except ImportError:
    print("需要安装openpyxl: pip install openpyxl")
    sys.exit(1)

excel_path = "/home/ubuntu/workspace/data/电机电力电气计算表.xlsx"

if not os.path.exists(excel_path):
    print(f"文件不存在: {excel_path}")
    sys.exit(1)

print("=" * 60)
print("Excel文件分析报告")
print("=" * 60)

wb = load_workbook(excel_path, data_only=False)

print(f"\n工作表数量: {len(wb.sheetnames)}")
print(f"工作表列表: {wb.sheetnames}")

# 分析第一个工作表
if wb.sheetnames:
    first_sheet_name = wb.sheetnames[0]
    ws = wb[first_sheet_name]
    
    print(f"\n{'='*60}")
    print(f"第一个工作表: {first_sheet_name}")
    print(f"{'='*60}")
    
    # 获取工作表维度
    print(f"使用范围: {ws.dimensions}")
    
    # 分析前20行，查找输入和输出区域
    print("\n前20行内容预览:")
    print("-" * 60)
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=False), 1):
        row_data = []
        for cell in row[:10]:  # 只显示前10列
            if cell.value is not None:
                if cell.data_type == 'f':  # 公式
                    row_data.append(f"{cell.coordinate}:{cell.value[:50] if len(str(cell.value)) > 50 else cell.value}")
                else:
                    row_data.append(f"{cell.coordinate}:{str(cell.value)[:30]}")
        if row_data:
            print(f"行{row_idx}: {' | '.join(row_data)}")
    
    # 查找包含公式的单元格
    formulas = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == 'f' and cell.value:
                formulas.append({
                    'cell': cell.coordinate,
                    'formula': cell.value,
                    'value': cell.value if cell.data_type != 'f' else None
                })
    
    if formulas:
        print(f"\n找到 {len(formulas)} 个公式单元格:")
        print("-" * 60)
        for i, f in enumerate(formulas[:10], 1):  # 只显示前10个
            print(f"{i}. {f['cell']}: {f['formula'][:80]}")
        if len(formulas) > 10:
            print(f"... 还有 {len(formulas) - 10} 个公式")
    
    # 查找可能的输入单元格（通常是文本标签旁边的空单元格或数值单元格）
    print("\n可能的输入区域分析:")
    print("-" * 60)
    input_candidates = []
    for row in ws.iter_rows(min_row=1, max_row=30):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                # 查找包含"输入"、"参数"、"值"等关键词的单元格
                if any(keyword in str(cell.value) for keyword in ['输入', '参数', '值', 'Input', 'Parameter', 'Value']):
                    # 检查右侧或下方的单元格
                    right_cell = ws.cell(cell.row, cell.column + 1)
                    if right_cell.value is None or isinstance(right_cell.value, (int, float)):
                        input_candidates.append({
                            'label': cell.value,
                            'input_cell': right_cell.coordinate,
                            'position': f"行{cell.row}, 列{cell.column}"
                        })
    
    if input_candidates:
        for candidate in input_candidates[:10]:
            print(f"标签: {candidate['label']} -> 输入单元格: {candidate['input_cell']} ({candidate['position']})")
    
    print("\n" + "=" * 60)

wb.close()

