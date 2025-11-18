#!/usr/bin/env python3
"""
分析"不同形状物体惯量计算"工作表
"""
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from openpyxl import load_workbook

EXCEL_PATH = "/home/ubuntu/workspace/data/电机电力电气计算表.xlsx"

def main():
    wb = load_workbook(EXCEL_PATH, data_only=False)
    ws = wb["不同形状物体惯量计算"]
    
    print("="*80)
    print("不同形状物体惯量计算 - 工作表检查报告")
    print("="*80)
    
    print(f"\n【基本信息】")
    print(f"工作表范围: {ws.dimensions}")
    print(f"总行数: {ws.max_row}")
    print(f"总列数: {ws.max_column}")
    
    # 识别所有计算模块
    print(f"\n【计算模块识别】")
    print("-"*80)
    
    modules = [
        {
            "name": "1. 圆柱体惯量计算（平行）",
            "start_row": 11,
            "end_row": 24,
            "description": "圆柱体长度方向中心线和旋转中心线平行",
            "inputs": ["外径d0(mm)", "内径d1(mm)", "长度L(mm)", "密度ρ(kg/m³)", "重心线与旋转轴线距离e(mm)"],
            "outputs": ["物体质量m(kg)", "物体惯量(kg·cm²)"]
        },
        {
            "name": "2. 圆柱体惯量计算（垂直）",
            "start_row": 26,
            "end_row": 39,
            "description": "圆柱体长度方向中心线和旋转中心线垂直",
            "inputs": ["外径d0(mm)", "内径d1(mm)", "长度L(mm)", "密度ρ(kg/m³)", "重心线与旋转轴线距离e(mm)"],
            "outputs": ["物体质量m(kg)", "物体惯量(kg·cm²)"]
        },
        {
            "name": "3. 方形物体惯量计算",
            "start_row": 41,
            "end_row": 54,
            "description": "长方体（方形物体）惯量计算",
            "inputs": ["长度x(mm)", "宽度y(mm)", "高度z(mm)", "密度ρ(kg/m³)", "重心线与旋转轴线距离e(mm)"],
            "outputs": ["物体质量m(kg)", "物体惯量(kg·cm²)"]
        },
        {
            "name": "4. 饼状物体惯量计算",
            "start_row": 56,
            "end_row": 68,
            "description": "实心圆柱体（饼状）惯量计算",
            "inputs": ["直径d(mm)", "厚度h(mm)", "密度ρ(kg/m³)", "重心线与旋转轴线距离e(mm)"],
            "outputs": ["物体质量m(kg)", "物体惯量(kg·cm²)"]
        },
        {
            "name": "5. 直线运动物体惯量计算",
            "start_row": 70,
            "end_row": 77,
            "description": "将直线运动转换为旋转惯量",
            "inputs": ["电机每转1圈物体直线运动量A(mm)", "物体质量m(kg)"],
            "outputs": ["物体惯量(kg·cm²)"]
        },
        {
            "name": "6. 直接惯量计算",
            "start_row": 79,
            "end_row": 91,
            "description": "已知惯量J0和质量m，计算平移后的惯量J1",
            "inputs": ["惯量J0(kg·cm²)", "质量m(kg)", "重心线与旋转轴线距离e(mm)"],
            "outputs": ["质量m1(kg)", "惯量J1(kg·cm²)"]
        }
    ]
    
    for i, module in enumerate(modules, 1):
        print(f"\n{module['name']}")
        print(f"  位置: 行{module['start_row']}-{module['end_row']}")
        print(f"  说明: {module['description']}")
        print(f"  输入参数: {', '.join(module['inputs'])}")
        print(f"  输出结果: {', '.join(module['outputs'])}")
    
    # 分析所有公式
    print(f"\n【公式统计】")
    print("-"*80)
    
    formulas = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == 'f' and cell.value:
                formulas.append({
                    'cell': cell.coordinate,
                    'formula': cell.value,
                    'row': cell.row
                })
    
    print(f"总公式数: {len(formulas)}")
    
    # 按模块分类公式
    print(f"\n【各模块公式】")
    print("-"*80)
    
    for module in modules:
        module_formulas = [f for f in formulas if module['start_row'] <= f['row'] <= module['end_row']]
        print(f"\n{module['name']} (行{module['start_row']}-{module['end_row']}):")
        for f in module_formulas:
            print(f"  {f['cell']:8s}: {f['formula']}")
    
    # 材料密度参考
    print(f"\n【材料密度参考】")
    print("-"*80)
    print("铁: 7.9×10³ kg/m³")
    print("铝: 2.8×10³ kg/m³")
    print("黄铜: 8.5×10³ kg/m³")
    print("尼龙: 1.1×10³ kg/m³")
    print("π = 3.14159")
    
    # 单位换算
    print(f"\n【单位换算】")
    print("-"*80)
    print("长度: mm → m (除以1000)")
    print("惯量: kg·cm² → kg·m² (除以10000)")
    print("距离: mm → m (除以1000 或 除以100，取决于公式)")
    
    print(f"\n" + "="*80)
    print("检查完成，等待用户逐一分析")
    print("="*80)
    
    wb.close()

if __name__ == "__main__":
    main()

