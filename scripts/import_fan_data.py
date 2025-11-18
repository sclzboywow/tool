"""
从Excel导入风机性能数据到SQLite数据库
"""
import sys
import os
from pathlib import Path

# 添加项目根目录到路径
BASE_DIR = Path(__file__).parent.parent
sys.path.insert(0, str(BASE_DIR))

from openpyxl import load_workbook
from app.db.database import init_db, insert_fan_performance


def import_from_excel(excel_path: str, fan_type: str, sheet_name: str = None):
    """
    从Excel文件导入风机性能数据
    
    Args:
        excel_path: Excel文件路径
        fan_type: 风机型号，如 "4-68"
        sheet_name: 工作表名称，如果为None则使用第一个工作表
    """
    print(f"正在从 {excel_path} 导入 {fan_type} 型号的数据...")
    
    wb = load_workbook(excel_path, data_only=True)
    
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    
    print(f"使用工作表: {ws.title}")
    print(f"行数: {ws.max_row}, 列数: {ws.max_column}")
    
    # 这里需要根据实际的Excel结构来解析数据
    # 示例：假设数据在特定位置
    # 实际使用时需要根据Excel文件的具体结构调整
    
    print("注意：需要根据实际Excel文件结构来解析数据")
    print("请修改此脚本以匹配您的Excel文件格式")
    
    wb.close()


def import_from_calculated_data():
    """
    从计算结果反推的数据导入（4-68型号）
    这些数据是从"风机选型计算"工作表的计算结果反推出来的
    """
    print("导入4-68型号的性能数据（从计算结果反推）...")
    
    # 4-68型号的性能数据（从计算结果反推）
    fan_type = "4-68"
    performance_data = [
        {"point_index": 1, "phi": 0.165, "psi_p": 0.498073, "eta": 87.6},
        {"point_index": 2, "phi": 0.185, "psi_p": 0.487093, "eta": 90.3},
        {"point_index": 3, "phi": 0.205, "psi_p": 0.472080, "eta": 92.2},
        {"point_index": 4, "phi": 0.225, "psi_p": 0.450084, "eta": 93.0},
        {"point_index": 5, "phi": 0.245, "psi_p": 0.422075, "eta": 92.0},
        {"point_index": 6, "phi": 0.265, "psi_p": 0.388054, "eta": 88.5},
        {"point_index": 7, "phi": 0.285, "psi_p": 0.350073, "eta": 84.7},
    ]
    
    for point in performance_data:
        insert_fan_performance(
            fan_type=fan_type,
            point_index=point["point_index"],
            phi=point["phi"],
            psi_p=point["psi_p"],
            eta=point["eta"]
        )
        print(f"  导入点{point['point_index']}: φ={point['phi']}, ψ_p={point['psi_p']}, η={point['eta']}%")
    
    print(f"✓ 成功导入 {len(performance_data)} 个性能点")


def import_from_builtin_data():
    """
    从代码中内置的数据导入（如果与反推数据不同）
    """
    print("导入内置性能数据...")
    
    # 如果代码中有其他型号的数据，可以在这里添加
    pass


if __name__ == "__main__":
    # 初始化数据库
    print("=" * 80)
    print("风机性能数据导入工具")
    print("=" * 80)
    
    init_db()
    
    # 导入反推数据
    import_from_calculated_data()
    
    # 如果有Excel文件，可以导入
    # excel_path = "data/电机电力电气计算表.xlsx"
    # if os.path.exists(excel_path):
    #     import_from_excel(excel_path, "4-68", "风机选型计算 ")
    
    print("\n" + "=" * 80)
    print("数据导入完成！")
    print("=" * 80)

