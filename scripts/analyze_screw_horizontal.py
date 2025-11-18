#!/usr/bin/env python3
"""
分析"丝杠水平运动选型计算"工作表
"""
from openpyxl import load_workbook

EXCEL_PATH = "/home/ubuntu/workspace/data/电机电力电气计算表.xlsx"

def main():
    wb = load_workbook(EXCEL_PATH, data_only=False)
    ws = wb["丝杠水平运动 "]
    
    print("="*80)
    print("丝杠水平运动选型计算 - 工作表分析报告")
    print("="*80)
    
    print("\n【输入参数】")
    print("-"*80)
    inputs = [
        ("Vl", "速度", "m/min", "F4"),
        ("M", "滑动部分质量", "kg", "F5"),
        ("LB", "丝杠长度", "m", "F6"),
        ("DB", "丝杠直径", "m", "F7"),
        ("PB", "丝杠导程", "m", "F8"),
        ("MC", "连轴器质量", "kg", "F9"),
        ("DC", "连轴器直径", "m", "F10"),
        ("μ", "摩擦系数", "", "F11"),
        ("L", "移动距离", "m", "F12"),
        ("η", "机械效率", "", "F13"),
        ("t", "定位时间", "s", "F14"),
        ("A", "加减速时间比", "", "F15"),
        ("FA", "外力", "N", "F16"),
        ("a", "移动方向与水平轴夹角", "°", "F17"),
    ]
    
    for var, name, unit, cell in inputs:
        value = ws[cell].value
        print(f"{var:3s} ({name:15s}): {value} {unit} - {cell}")
    
    print("\n【常数】")
    print("-"*80)
    constants = [
        ("G", "重力加速度", "9.8", "m/s²", "K5"),
        ("π", "圆周率", "3.1416", "", "K6"),
        ("ρ", "丝杠密度", "7900", "kg/m³", "K7"),
        ("S", "安全系数", "1", "", "K45"),
        ("JM", "电机惯量", "0.0011", "kg·m²", "K49"),
        ("i", "减速机减速比", "4", "", "K62"),
    ]
    
    for var, name, value, unit, cell in constants:
        print(f"{var:3s} ({name:15s}): {value} {unit} - {cell}")
    
    print("\n【计算步骤和公式】")
    print("-"*80)
    
    calculations = [
        {
            "step": "1) 速度曲线",
            "name": "加速时间",
            "var": "t0",
            "formula": "t0 = t × A",
            "excel": "F21 = F14*F15",
            "unit": "s"
        },
        {
            "step": "2) 电机转速",
            "name": "电机转速",
            "var": "NM",
            "formula": "NM = Vl / PB",
            "excel": "F24 = F4/F8",
            "unit": "rpm"
        },
        {
            "step": "3) 负荷转矩计算",
            "name": "轴向负载",
            "var": "F",
            "formula": "F = FA + M×G×(sin(a) + μ×cos(a))",
            "excel": "F27 = F16+F5*K5*(SIN((F17/360)*2*K6)+F11*COS((F17/360)*2*K6))",
            "unit": "N"
        },
        {
            "step": "3) 负荷转矩计算",
            "name": "负载转矩",
            "var": "TL",
            "formula": "TL = (F × PB) / (2π × η)",
            "excel": "F30 = (F27*F8)/(2*K6*F13)",
            "unit": "Nm"
        },
        {
            "step": "4) 克服惯量的加速转矩计算",
            "name": "直线运动平台与负载惯量",
            "var": "JL1",
            "formula": "JL1 = M × (PB/(2π))²",
            "excel": "F34 = F5*(F8/(2*K6))^2",
            "unit": "kg·m²"
        },
        {
            "step": "4) 克服惯量的加速转矩计算",
            "name": "滚珠丝杠惯量",
            "var": "JB",
            "formula": "JB = π × ρ × LB × DB⁴ / 32",
            "excel": "F37 = K6*K7*F6*F7^4/32",
            "unit": "kg·m²"
        },
        {
            "step": "4) 克服惯量的加速转矩计算",
            "name": "连轴器惯量",
            "var": "JC",
            "formula": "JC = MC × DC² / 8",
            "excel": "F40 = F9*F10^2/8",
            "unit": "kg·m²"
        },
        {
            "step": "4) 克服惯量的加速转矩计算",
            "name": "总负荷惯量",
            "var": "JL",
            "formula": "JL = JL1 + JB + JC",
            "excel": "F43 = F34+F37+F40",
            "unit": "kg·m²"
        },
        {
            "step": "4) 克服惯量的加速转矩计算",
            "name": "启动转矩",
            "var": "TS",
            "formula": "TS = 2π × NM × (JM + JL) / (60 × t0)",
            "excel": "F46 = 2*K6*F24*(K49+F43)/(60*F21)",
            "unit": "Nm"
        },
        {
            "step": "5) 必须转矩",
            "name": "必须转矩",
            "var": "TM",
            "formula": "TM = (TL + TS) × S",
            "excel": "F50 = (F30+F46)*K45",
            "unit": "Nm"
        },
        {
            "step": "7) 负荷与电机惯量比",
            "name": "惯量比",
            "var": "I1",
            "formula": "I1 = JL / JM",
            "excel": "F57 = F43/K49",
            "unit": ""
        },
        {
            "step": "8) 负荷与减速机惯量比",
            "name": "折算后的惯量比",
            "var": "I2",
            "formula": "I2 = JL / (JM × i²)",
            "excel": "F62 = F43/(K49*K62^2)",
            "unit": ""
        },
    ]
    
    for calc in calculations:
        print(f"\n{calc['step']} - {calc['name']} ({calc['var']})")
        print(f"  公式: {calc['formula']}")
        print(f"  Excel: {calc['excel']}")
        print(f"  单位: {calc['unit']}")
    
    print("\n" + "="*80)
    print("分析完成")
    print("="*80)
    
    wb.close()

if __name__ == "__main__":
    main()

