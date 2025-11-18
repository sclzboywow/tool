#!/usr/bin/env python3
"""
Excel文件查看工具 - 用于AI理解Excel表格结构
可以查看指定工作表的内容、公式、格式等
"""
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json

EXCEL_PATH = "/home/ubuntu/workspace/data/电机电力电气计算表.xlsx"

def view_sheet(workbook, sheet_name, max_rows=50, max_cols=15):
    """查看指定工作表的内容"""
    if sheet_name not in workbook.sheetnames:
        print(f"错误：工作表 '{sheet_name}' 不存在")
        print(f"可用工作表：{workbook.sheetnames}")
        return
    
    ws = workbook[sheet_name]
    print(f"\n{'='*80}")
    print(f"工作表: {sheet_name}")
    print(f"使用范围: {ws.dimensions}")
    print(f"{'='*80}\n")
    
    # 获取实际使用的范围
    if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None:
        print("工作表为空")
        return
    
    # 显示表头（列号）
    header = "行号 | "
    for col in range(1, min(ws.max_column + 1, max_cols + 1)):
        header += f"{get_column_letter(col):>8} | "
    print(header)
    print("-" * len(header))
    
    # 显示数据
    for row_idx in range(1, min(ws.max_row + 1, max_rows + 1)):
        row_data = f"{row_idx:4d} | "
        for col_idx in range(1, min(ws.max_column + 1, max_cols + 1)):
            cell = ws.cell(row_idx, col_idx)
            cell_value = ""
            
            if cell.value is not None:
                if cell.data_type == 'f':  # 公式
                    # 显示公式
                    formula = str(cell.value)
                    if len(formula) > 8:
                        formula = formula[:5] + "..."
                    cell_value = f"={formula}"
                elif isinstance(cell.value, (int, float)):
                    cell_value = f"{cell.value:.2f}" if isinstance(cell.value, float) else str(cell.value)
                else:
                    cell_value = str(cell.value)[:8]
            
            # 添加格式信息（简化处理，避免错误）
            try:
                if cell.font.bold:
                    cell_value = "**" + cell_value
            except:
                pass
            
            row_data += f"{cell_value:>8} | "
        
        print(row_data)
    
    if ws.max_row > max_rows:
        print(f"... (还有 {ws.max_row - max_rows} 行未显示)")
    if ws.max_column > max_cols:
        print(f"... (还有 {ws.max_column - max_cols} 列未显示)")

def list_sheets(workbook):
    """列出所有工作表"""
    print(f"\n{'='*80}")
    print("Excel文件中的所有工作表：")
    print(f"{'='*80}")
    for i, sheet_name in enumerate(workbook.sheetnames, 1):
        ws = workbook[sheet_name]
        print(f"{i}. {sheet_name} (范围: {ws.dimensions})")
    print()

def get_formulas(workbook, sheet_name):
    """获取指定工作表的所有公式"""
    if sheet_name not in workbook.sheetnames:
        print(f"错误：工作表 '{sheet_name}' 不存在")
        return
    
    ws = workbook[sheet_name]
    formulas = []
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == 'f' and cell.value:
                formulas.append({
                    'cell': cell.coordinate,
                    'formula': cell.value,
                    'row': cell.row,
                    'col': cell.column
                })
    
    if formulas:
        print(f"\n{'='*80}")
        print(f"工作表 '{sheet_name}' 中的公式 ({len(formulas)} 个):")
        print(f"{'='*80}\n")
        for f in formulas:
            print(f"{f['cell']:8s} (行{f['row']:3d}, 列{f['col']:2d}): {f['formula']}")
    else:
        print(f"工作表 '{sheet_name}' 中没有找到公式")

def get_cell_range(workbook, sheet_name, range_str):
    """获取指定单元格范围的内容"""
    if sheet_name not in workbook.sheetnames:
        print(f"错误：工作表 '{sheet_name}' 不存在")
        return
    
    ws = workbook[sheet_name]
    
    try:
        cells = ws[range_str]
        print(f"\n{'='*80}")
        print(f"工作表 '{sheet_name}' 范围 '{range_str}':")
        print(f"{'='*80}\n")
        
        # 处理单个单元格
        if not isinstance(cells, tuple):
            cell = cells
            print(f"单元格 {cell.coordinate}:")
            print(f"  值: {cell.value}")
            if cell.data_type == 'f':
                print(f"  公式: {cell.value}")
            print(f"  数据类型: {cell.data_type}")
            return
        
        # 处理单元格范围
        for row in cells:
            row_data = []
            for cell in row:
                cell_info = f"{cell.coordinate}:"
                if cell.value is not None:
                    if cell.data_type == 'f':
                        cell_info += f"={cell.value}"
                    else:
                        cell_info += str(cell.value)
                else:
                    cell_info += "(空)"
                row_data.append(cell_info)
            print(" | ".join(row_data))
    except Exception as e:
        print(f"错误：无法读取范围 '{range_str}': {e}")

def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"错误：Excel文件不存在: {EXCEL_PATH}")
        sys.exit(1)
    
    print(f"正在加载Excel文件: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH, data_only=False)  # data_only=False 保留公式
    
    if len(sys.argv) == 1:
        # 没有参数，显示所有工作表列表
        list_sheets(wb)
        print("\n使用方法：")
        print("  python3 view_excel.py <工作表名>              # 查看整个工作表")
        print("  python3 view_excel.py <工作表名> <范围>       # 查看指定范围（如 A1:C10）")
        print("  python3 view_excel.py <工作表名> --formulas   # 查看所有公式")
        print("  python3 view_excel.py --list                  # 列出所有工作表")
        wb.close()
        return
    
    if sys.argv[1] == '--list':
        list_sheets(wb)
        wb.close()
        return
    
    sheet_name = sys.argv[1]
    
    if len(sys.argv) == 2:
        # 只指定了工作表名，查看整个工作表
        view_sheet(wb, sheet_name)
    elif sys.argv[2] == '--formulas':
        # 查看所有公式
        get_formulas(wb, sheet_name)
    else:
        # 查看指定范围
        range_str = sys.argv[2]
        get_cell_range(wb, sheet_name, range_str)
    
    wb.close()

if __name__ == "__main__":
    main()

