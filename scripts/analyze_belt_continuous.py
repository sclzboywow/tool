#!/usr/bin/env python3
"""
分析"皮带轮连续运动"工作表
"""
from openpyxl import load_workbook

EXCEL_PATH = "/home/ubuntu/workspace/data/电机电力电气计算表.xlsx"

def main():
    try:
        wb = load_workbook(EXCEL_PATH, data_only=False)
        
        # 尝试查找连续运动相关的工作表
        sheet_names = wb.sheetnames
        print("="*80)
        print("Excel工作表列表")
        print("="*80)
        for i, name in enumerate(sheet_names, 1):
            print(f"{i}. {name}")
        
        # 查找可能的工作表名称
        possible_names = ["皮带轮连续运动", "皮带连续运动", "连续运动", "belt continuous", "continuous"]
        ws = None
        ws_name = None
        
        for name in possible_names:
            if name in sheet_names:
                ws = wb[name]
                ws_name = name
                break
        
        # 如果没找到，尝试查找包含"连续"或"continuous"的工作表
        if ws is None:
            for name in sheet_names:
                if "连续" in name or "continuous" in name.lower():
                    ws = wb[name]
                    ws_name = name
                    break
        
        if ws is None:
            print("\n未找到'皮带轮连续运动'相关的工作表")
            print("\n请检查Excel文件中是否有以下名称的工作表：")
            for name in possible_names:
                print(f"  - {name}")
            return
        
        print(f"\n找到工作表: {ws_name}")
        print("="*80)
        print("皮带轮连续运动选型计算 - 工作表分析报告")
        print("="*80)
        
        # 打印前100行，查找关键信息
        print("\n【工作表内容预览】")
        print("-"*80)
        for row in range(1, min(100, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(20, ws.max_column + 1)):
                cell = ws.cell(row, col)
                if cell.value:
                    value_str = str(cell.value)[:40]
                    row_data.append(f"{cell.coordinate}:{value_str}")
            if row_data:
                print(f"行{row}: {' | '.join(row_data)}")
        
        print("\n【查找公式】")
        print("-"*80)
        formulas_found = []
        for row in range(1, min(150, ws.max_row + 1)):
            for col in range(1, min(25, ws.max_column + 1)):
                cell = ws.cell(row, col)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formulas_found.append((cell.coordinate, cell.value))
                    print(f"{cell.coordinate}: {cell.value}")
        
        if not formulas_found:
            print("未找到公式（可能工作表使用data_only=True读取）")
            print("\n尝试读取计算值...")
            wb_data = load_workbook(EXCEL_PATH, data_only=True)
            ws_data = wb_data[ws_name]
            for row in range(1, min(100, ws_data.max_row + 1)):
                for col in range(1, min(20, ws_data.max_column + 1)):
                    cell = ws_data.cell(row, col)
                    if cell.value is not None:
                        print(f"{cell.coordinate}: {cell.value}")
        
        print("\n【查找关键参数】")
        print("-"*80)
        # 查找可能的参数标签
        param_keywords = ["V", "mL", "m2", "D", "μ", "η", "ηG", "i", "FA", "a", "S", "JM", 
                         "线速度", "质量", "直径", "摩擦", "效率", "减速比", "外力", "夹角", "安全系数", "惯量"]
        for row in range(1, min(100, ws.max_row + 1)):
            for col in range(1, min(20, ws.max_column + 1)):
                cell = ws.cell(row, col)
                if cell.value:
                    cell_str = str(cell.value)
                    for keyword in param_keywords:
                        if keyword in cell_str:
                            print(f"{cell.coordinate}: {cell_str}")
                            break
        
    except FileNotFoundError:
        print(f"错误: 找不到文件 {EXCEL_PATH}")
    except Exception as e:
        print(f"错误: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()

